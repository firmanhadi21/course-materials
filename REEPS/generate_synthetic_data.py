"""
Generate synthetic REEPS observation data for workshop use.
Randomizes locations and observation counts while maintaining
the same species, survey methods, and general spatial/temporal patterns.
Output format matches the real REEPS_Biodiversity_Database.xlsx structure.
"""
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import random

np.random.seed(None)  # truly random each run

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── REEPS species with realistic relative abundance weights ──────────
REEPS_SPECIES = [
    {'sci': 'Panthera pardus melas',       'common': 'Javan Leopard',           'iucn': 'CR', 'weight': 0.13},
    {'sci': 'Nycticebus javanicus',        'common': 'Javan Slow Loris',        'iucn': 'CR', 'weight': 0.19},
    {'sci': 'Manis javanica',              'common': 'Sunda Pangolin',          'iucn': 'CR', 'weight': 0.10},
    {'sci': 'Hylobates moloch',            'common': 'Javan Gibbon',            'iucn': 'EN', 'weight': 0.12},
    {'sci': 'Presbytis comata',            'common': 'Grizzled Langur',         'iucn': 'EN', 'weight': 0.08},
    {'sci': 'Trachypithecus auratus',      'common': 'Javan Langur',            'iucn': 'VU', 'weight': 0.12},
    {'sci': 'Nisaetus bartelsi',           'common': "Bartels's Hawk-eagle",    'iucn': 'VU', 'weight': 0.01},
    {'sci': 'Aonyx cinereus',              'common': 'Asian Small-clawed Otter','iucn': 'VU', 'weight': 0.09},
    {'sci': 'Tragulus kanchil',            'common': 'Lesser Mouse-deer',       'iucn': 'LC', 'weight': 0.02},
    {'sci': 'Prionailurus bengalensis',    'common': 'Leopard Cat',             'iucn': 'LC', 'weight': 0.10},
    {'sci': 'Paradoxurus hermaphroditus',  'common': 'Common Palm Civet',       'iucn': 'LC', 'weight': 0.04},
]

# Non-REEPS bycatch species (smaller counts)
NON_REEPS = [
    {'sci': 'Macaca fascicularis',    'common': 'Long-tailed Macaque'},
    {'sci': 'Sus scrofa',             'common': 'Wild Boar'},
    {'sci': 'Hystrix javanica',       'common': 'Javan Porcupine'},
    {'sci': 'Paguma larvata',         'common': 'Masked Palm Civet'},
]

SURVEY_METHODS = ['Observation', 'Camera Trap', 'Sign (Feces)', 'Sign (Track)',
                  'Interview', 'Sign (Nest)', 'Unknown']
METHOD_WEIGHTS = [0.30, 0.10, 0.15, 0.10, 0.15, 0.05, 0.15]

# Survey periods with effort weights (more records in recent years)
SURVEY_YEARS = [
    (2009, 0.002), (2012, 0.02), (2014, 0.06), (2017, 0.25),
    (2018, 0.005), (2020, 0.25), (2022, 0.20), (2024, 0.10),
    (2025, 0.07), (2026, 0.04),
]

# AOI bounding box (approximate UCPS area, slightly jittered)
LAT_CENTER, LAT_SPREAD = -6.955, 0.025
LON_CENTER, LON_SPREAD = 107.225, 0.045

# Spatial hotspots (3 clusters where observations concentrate)
HOTSPOTS = [
    (-6.945, 107.222, 0.008),   # central-eastern richness hotspot
    (-6.955, 107.210, 0.012),   # western corridor
    (-6.965, 107.235, 0.010),   # southern patch
]

SOURCES = {
    2009: 'Satwa Target 2020', 2012: 'Satwa Target 2020',
    2014: 'Temuan Langsung', 2017: 'Satwa Target 2020',
    2018: 'Temuan Langsung', 2020: 'Data Gabungan',
    2022: 'Species_coord_2022', 2024: 'reeps_record_24',
    2025: 'BMP 2025-2026', 2026: 'BMP 2025-2026',
}

MONTHS_BY_YEAR = {
    2025: [6, 7, 8, 9, 10, 11, 12],
    2026: [1, 2],
}


def random_location():
    """Generate a random location, clustered around hotspots."""
    if random.random() < 0.7:
        # 70% near a hotspot
        h = random.choice(HOTSPOTS)
        lat = np.random.normal(h[0], h[2])
        lon = np.random.normal(h[1], h[2])
    else:
        # 30% anywhere in AOI
        lat = np.random.normal(LAT_CENTER, LAT_SPREAD)
        lon = np.random.normal(LON_CENTER, LON_SPREAD)
    # Clamp to AOI
    lat = np.clip(lat, -6.985, -6.925)
    lon = np.clip(lon, 107.170, 107.280)
    return round(lat, 6), round(lon, 6)


def weighted_choice(items, weights):
    return random.choices(items, weights=weights, k=1)[0]


def generate_records(n_reeps=480, n_bycatch=80):
    """Generate synthetic observation records."""
    records = []
    no = 1

    # REEPS records
    sp_weights = [s['weight'] for s in REEPS_SPECIES]
    year_vals = [y for y, _ in SURVEY_YEARS]
    year_weights = [w for _, w in SURVEY_YEARS]

    for _ in range(n_reeps):
        sp = weighted_choice(REEPS_SPECIES, sp_weights)
        year = weighted_choice(year_vals, year_weights)
        method = weighted_choice(SURVEY_METHODS, METHOD_WEIGHTS)
        lat, lon = random_location()

        month = None
        if year in MONTHS_BY_YEAR:
            month = random.choice(MONTHS_BY_YEAR[year])

        month_year = f'{year}'
        if month:
            import calendar
            month_year = f'{calendar.month_abbr[month]}-{year}'

        records.append({
            'No': no,
            'Source': SOURCES.get(year, 'Unknown'),
            'Species': sp['sci'],
            'Common_Name': sp['common'],
            'Status': 'REEPS',
            'Year': year,
            'Month': month,
            'Month_Year': month_year,
            'Survey_Method': method,
            'Location': None,
            'Latitude': lat,
            'Longitude': lon,
        })
        no += 1

    # Non-REEPS bycatch
    for _ in range(n_bycatch):
        sp = random.choice(NON_REEPS)
        year = weighted_choice(year_vals, year_weights)
        method = weighted_choice(SURVEY_METHODS, METHOD_WEIGHTS)
        lat, lon = random_location()

        records.append({
            'No': no,
            'Source': SOURCES.get(year, 'Unknown'),
            'Species': sp['sci'],
            'Common_Name': sp['common'],
            'Status': None,
            'Year': year,
            'Month': None,
            'Month_Year': str(year),
            'Survey_Method': method,
            'Location': None,
            'Latitude': lat,
            'Longitude': lon,
        })
        no += 1

    # Sort by year then species
    records.sort(key=lambda r: (r['Year'], r['Species']))
    for i, r in enumerate(records):
        r['No'] = i + 1

    return records


def write_excel(records, path):
    """Write records to Excel in the same format as REEPS_Biodiversity_Database.xlsx."""
    wb = openpyxl.Workbook()

    # Master Database sheet
    ws = wb.active
    ws.title = 'Master Database'
    ws.cell(1, 1, 'BIODIVERSITY SURVEY – MASTER DATABASE (SYNTHETIC / WORKSHOP)')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, 'Synthetic data for workshop use — randomized locations and counts')
    ws.cell(2, 1).font = Font(italic=True)

    headers = ['No', 'Source', 'Species', 'Common Name', 'Status', 'Year', 'Month',
               'Month-Year', 'Survey Method', 'Location', 'Latitude', 'Longitude']
    for j, h in enumerate(headers, 1):
        ws.cell(4, j, h)
        ws.cell(4, j).font = Font(bold=True)

    for i, r in enumerate(records):
        row = i + 5
        ws.cell(row, 1, r['No'])
        ws.cell(row, 2, r['Source'])
        ws.cell(row, 3, r['Species'])
        ws.cell(row, 4, r['Common_Name'])
        ws.cell(row, 5, r['Status'])
        ws.cell(row, 6, r['Year'])
        ws.cell(row, 7, r['Month'])
        ws.cell(row, 8, r['Month_Year'])
        ws.cell(row, 9, r['Survey_Method'])
        ws.cell(row, 10, r['Location'])
        ws.cell(row, 11, r['Latitude'])
        ws.cell(row, 12, r['Longitude'])

    # REEPS Database sheet (filtered)
    ws_r = wb.create_sheet('REEPS Database')
    ws_r.cell(1, 1, 'REEPS SPECIES DATABASE (SYNTHETIC / WORKSHOP)')
    ws_r.cell(1, 1).font = Font(bold=True, size=14)

    reeps_records = [r for r in records if r['Status'] == 'REEPS']
    ws_r.cell(2, 1, f'Synthetic REEPS data | {len(reeps_records)} records | 11 species')

    for j, h in enumerate(headers, 1):
        ws_r.cell(4, j, h)
        ws_r.cell(4, j).font = Font(bold=True)

    for i, r in enumerate(reeps_records):
        row = i + 5
        ws_r.cell(row, 1, i + 1)
        ws_r.cell(row, 2, r['Source'])
        ws_r.cell(row, 3, r['Species'])
        ws_r.cell(row, 4, r['Common_Name'])
        ws_r.cell(row, 5, r['Status'])
        ws_r.cell(row, 6, r['Year'])
        ws_r.cell(row, 7, r['Month'])
        ws_r.cell(row, 8, r['Month_Year'])
        ws_r.cell(row, 9, r['Survey_Method'])
        ws_r.cell(row, 10, r['Location'])
        ws_r.cell(row, 11, r['Latitude'])
        ws_r.cell(row, 12, r['Longitude'])

    # Temporal Summary (empty placeholder)
    ws_t = wb.create_sheet('Temporal Summary')
    ws_t.cell(1, 1, 'REEPS SPECIES – TEMPORAL DISTRIBUTION')
    ws_t.cell(2, 1, 'To be populated by the analysis pipeline')

    wb.save(path)
    return len(records), len(reeps_records)


def write_csv(records, path):
    """Also write as CSV for easy inspection."""
    df = pd.DataFrame(records)
    df.to_csv(path, index=False)


def main():
    print("=" * 60)
    print("Generating Synthetic REEPS Dataset for Workshop")
    print("=" * 60)

    records = generate_records(n_reeps=480, n_bycatch=80)

    # Summary
    from collections import Counter
    sp_counts = Counter(r['Species'] for r in records if r['Status'] == 'REEPS')
    yr_counts = Counter(r['Year'] for r in records)

    print(f"\nTotal records: {len(records)}")
    print(f"REEPS records: {sum(1 for r in records if r['Status'] == 'REEPS')}")
    print(f"Non-REEPS: {sum(1 for r in records if r['Status'] != 'REEPS')}")
    print(f"\nSpecies breakdown:")
    for sp, n in sorted(sp_counts.items()):
        print(f"  {sp}: {n}")
    print(f"\nYear breakdown:")
    for y in sorted(yr_counts.keys()):
        print(f"  {y}: {yr_counts[y]}")

    # Write files
    xlsx_path = os.path.join(OUT_DIR, 'Workshop_Biodiversity_Database.xlsx')
    csv_path = os.path.join(OUT_DIR, 'workshop_synthetic_records.csv')

    n_total, n_reeps = write_excel(records, xlsx_path)
    write_csv(records, csv_path)

    print(f"\nOutput files:")
    print(f"  {xlsx_path} ({n_total} total, {n_reeps} REEPS)")
    print(f"  {csv_path}")
    print(f"\nThis synthetic data can now be processed through the full analysis pipeline.")


if __name__ == '__main__':
    main()
